#!python3

#creates multiplication table based on the argument provided (sys.argv)
import sys
import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
if len(sys.argv) > 1:
    num = int(sys.argv[1])
    sheet = wb.get_active_sheet()
    for i in range(2, num + 2):
        sheet['A' + str(i)] = i - 1
        sheet['A' + str(i)].font = Font(bold=True)
        sheet.cell(row=1, column=i).value = i - 1
        sheet.cell(row=1, column=i).font = Font(bold=True)

    counter = 2
    activator = True
    while activator == True:
        if counter > sheet.max_column:
            activator = False
            continue
        column_name = openpyxl.utils.get_column_letter(counter)
        for i in range(2, sheet.max_column + 1):
            sheet.cell(row=i, column=counter).value = '=A' + str(i) + '*' + column_name + str(1)
        counter += 1




wb.save('example.xlsx')
