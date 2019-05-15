
import openpyxl

wb = openpyxl.load_workbook('exampleo.xlsx')
sheet = wb.active

max_column = sheet.max_column
#print(max_column)
max_row = sheet.max_row

columns = []

for column in range(1, max_column + 1):
    data = []
    for cell in range(1, max_row + 1):
        cell_value = sheet.cell(row=cell, column=column).value
        #print(cell_value)
        data.append(cell_value)
    columns.append(data)


wb = openpyxl.Workbook()
sheet = wb.active

for row in range(0, len(columns)):
    for cell in range(0, len(columns[row])):
        sheet.cell(row=row +1, column=cell + 1).value = columns[row][cell]

wb.save('newfile.xlsx')