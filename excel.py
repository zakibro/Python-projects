"""converts .xlsx files into .csv - file a sheet"""
import os
import openpyxl
import csv

for excelFile in os.listdir('.'):
    if excelFile.endswith('.xlsx'):
        wb = openpyxl.load_workbook(excelFile)
        for sheetName in wb.get_sheet_names():
            sheet = wb.get_sheet_by_name(sheetName)

            excelFile = excelFile.rstrip('.xlsx')

            csvfile = open(excelFile + '_' + sheet.title + '.csv', 'w', newline='')
            csvWriter = csv.writer(csvfile)

            for rowNum in range(1, sheet.max_row + 1):
                rowData = []
                for colNum in range(1, sheet.max_column + 1):
                    cello = sheet.cell(row=rowNum, column=colNum).value
                    rowData.append(cello)
                csvWriter.writerow(rowData)

            csvfile.close()

